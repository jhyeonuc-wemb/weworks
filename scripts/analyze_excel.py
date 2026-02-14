#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 템플릿 분석 스크립트
수지분석서, 정산서 파일의 구조를 분석합니다.
"""

import os
import sys
from pathlib import Path

# 프로젝트 루트 경로 설정
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

def analyze_excel_file(file_path):
    """엑셀 파일 분석"""
    try:
        import openpyxl
        
        print(f"\n{'='*80}")
        print(f"파일: {os.path.basename(file_path)}")
        print(f"{'='*80}")
        
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        print(f"\n시트 목록 ({len(wb.sheetnames)}개):")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {sheet_name}")
        
        # 각 시트 분석
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n[시트: {sheet_name}]")
            print(f"  - 최대 행: {ws.max_row}")
            print(f"  - 최대 열: {ws.max_column}")
            
            # 첫 10행의 데이터 미리보기
            print(f"\n  첫 5행 미리보기:")
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=False), 1):
                row_values = []
                for cell in row[:10]:  # 처음 10개 셀만
                    if cell.value is not None:
                        if isinstance(cell.value, str):
                            value = cell.value[:30] if len(cell.value) > 30 else cell.value
                        else:
                            value = cell.value
                        # 수식인 경우 표시
                        if cell.data_type == 'f':
                            row_values.append(f"{value} [수식]")
                        else:
                            row_values.append(str(value))
                    else:
                        row_values.append("")
                print(f"    행 {row_idx}: {' | '.join(row_values)}")
            
            # 수식이 있는 셀 찾기
            formula_cells = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formula_cells.append({
                            'cell': cell.coordinate,
                            'formula': cell.value[:100]  # 처음 100자만
                        })
            if formula_cells:
                print(f"\n  수식이 있는 셀 ({len(formula_cells)}개, 최대 5개 표시):")
                for f in formula_cells[:5]:
                    print(f"    {f['cell']}: {f['formula']}")
        
        wb.close()
        return True
        
    except ImportError:
        print("openpyxl 라이브러리가 필요합니다. 설치: pip install openpyxl")
        return False
    except Exception as e:
        print(f"오류 발생: {e}")
        return False

def main():
    """메인 함수"""
    templates_dir = project_root / "docs" / "templates"
    
    if not templates_dir.exists():
        print(f"템플릿 폴더를 찾을 수 없습니다: {templates_dir}")
        return
    
    # 엑셀 파일 찾기
    excel_files = list(templates_dir.glob("*.xlsx")) + list(templates_dir.glob("*.xls"))
    
    if not excel_files:
        print("엑셀 파일을 찾을 수 없습니다.")
        return
    
    print(f"\n총 {len(excel_files)}개의 엑셀 파일을 찾았습니다.\n")
    
    # 파일별로 분석
    for excel_file in sorted(excel_files):
        analyze_excel_file(excel_file)
    
    print(f"\n{'='*80}")
    print("분석 완료!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    main()
